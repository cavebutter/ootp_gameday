from prettytable import PrettyTable
import sqlite3
import shutil
from pathlib import Path
from _datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import sqlite_depth_chart_functions as dp
from os import remove


######################################
#        WHAT THIS MODULE DOES       #
######################################
#                                    #
# 1. Imports small_db.py and creates #
#    a small database in memory      #
# 2. Has the user select League and  #
#    organization for the depth chart#
# 3. Copies excel template with new  #
#    new name to output              #
# 4. Retrieves db data for depth chart
# 5. Writes to Excel and Formats     #
######################################

###################
# Create small_db #
###################
import small_db

#######################
# Connect to small_db #
#######################
try:
    cnx = sqlite3.connect('small_db')
    cursor = cnx.cursor()
    print("Successfully Connected to small_db")


except sqlite3.Error as error:
    print("Error while creating a sqlite table", error)



##########################
# Select League and Team #
##########################
try:
    cursor.execute("SELECT league_id, name FROM leagues WHERE parent_league_id = 0")
except sqlite3.OperationalError:
    print("Exiting...")
    cnx.close()
    remove('small_db')
    exit()

league_list = cursor.fetchall()
valid_leagues = []
for pair in league_list:
    valid_leagues.append(str(pair[0]))

# - Make a Pretty Table to display ID and League
x = PrettyTable()
x.field_names = ['ID', 'League']
for item in league_list:
    x.add_row(item)
print(x)

# This bit to ensure that league_id is in league_list
league_id = ''
while league_id not in valid_leagues:
    if league_id == "quit" or league_id == "Quit":
        cnx.close()
        remove('small_db')
        exit()
    else:
        league_id = input("Enter the league of your choice (or 'Quit' to quit): ")



#  Get list of parent teams to select from
cursor.execute("""SELECT team_id, name || ' ' || nickname AS team
                    FROM teams
                    WHERE parent_team_id = 0 AND league_id=""" + str(league_id))
teams_list = cursor.fetchall()
valid_teams = []
for team in teams_list:
    valid_teams.append(str(team[0]))

# Make a Pretty Table
x = PrettyTable()
x.field_names = ['ID', 'Team']
for item in teams_list:
    x.add_row(item)
print(x)

#  Get org param
org_param = ''
while org_param not in valid_teams:
    if org_param == "quit" or org_param == "Quit":
        cnx.close()
        remove('small_db')
        exit()
    else:
        org_param = input("Enter the ID of the organization you want for your depth chart\n(Or 'quit' to exit) : ")

# Get org name
cursor.execute("SELECT name || ' ' || nickname FROM teams WHERE team_id=" + org_param)
result = cursor.fetchall()
team_name = result[0][0]

# Copy New xlsx with the Date and Org Name
str_date = datetime.strftime(small_db.game_date, '%m-%d-%Y')
new_file_name = team_name + "-" + str_date + '_depth_chart.xlsx'
src_file = Path.cwd() / 'depth_chart_template.xlsx'
dest_path = Path.cwd() / 'output' / new_file_name
shutil.copy(src_file, dest_path)

###########################################
#   Excel Formatting and Write Functions  #
###########################################

#  Change font color to white
#  white_text param is the cell range
def white_font(white_text):
    ft = Font(color='ffffff')
    for _row in ws[white_text]:
        for _cell in _row:
            _cell.font = ft

#  Borders and Alignment
def borders(top_left, bottom_right):
    cell_range = top_left + ":" + bottom_right
    for _row in ws[cell_range]:
        for _cell in _row:
            _cell.border = Border(top=Side(border_style="thin"),
                                  left=Side(border_style="thin"),
                                  right=Side(border_style="thin"),
                                  bottom=Side(border_style="thin")
                                  )
    for _row in ws[cell_range]:  #  Is this second loop necessary?  Can I just add it to the one above?
        for _cell in _row:
            _cell.alignment = Alignment(horizontal='center')

#############################
#  Write to Excel Functions #
#############################

def write_data(depth_chart):
    for i, line in enumerate(depth_chart):
        for k, val in enumerate(line):
            ws.cell(row=i + 2, column=k + 1).value = val

#  Open xlsx and find the correct tab
wb = openpyxl.load_workbook(dest_path)
ws = wb['depthchart_1B']

##############
# First Base #
##############
cursor.execute(dp.first_base + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_1b + dp.order_language)
first_base_depth = cursor.fetchall()

#  Write data beginning at Row 2

write_data(first_base_depth)

#  Apply white text formatting to Columns N - AG
row_limit = len(first_base_depth) + 2
white_text = "N2:AG" + str(row_limit)
white_font(white_text)

#  Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(first_base_depth)+1)
borders(top_left,bottom_right)

########################
#  Get OF Depth Chart  #
########################

cursor.execute(dp.outfield + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_of + dp.order_language)
of_depth = cursor.fetchall()

ws = wb['depthchart_of']

#  Write Data
write_data(of_depth)

#  Apply white text formatting to Columns N - AH
row_limit = len(of_depth) + 2
white_text = "N2:AH" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(of_depth)+1)
borders(top_left,bottom_right)

########################
#  Get C Depth Chart  #
########################

cursor.execute(dp.catcher + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_c + dp.order_language)
c_depth = cursor.fetchall()

ws = wb['depthchart_c']

#  Write Data
write_data(c_depth)

#  Apply white text formatting to Columns N - AF
row_limit = len(c_depth) + 2
white_text = "N2:AF" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(c_depth)+1)
borders(top_left,bottom_right)

########################
#  Get SP Depth Chart  #
########################

cursor.execute(dp.sp + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_sp + dp.order_language)
sp_depth = cursor.fetchall()

ws = wb['depthchart_sp']

#  Write Data
write_data(sp_depth)

#  Apply white text formatting to Columns O - X
row_limit = len(sp_depth) + 2
white_text = "O2:X" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "N" + str(len(sp_depth)+1)
borders(top_left,bottom_right)

########################
#  Get BP Depth Chart  #
########################

cursor.execute(dp.sp + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_bp + dp.order_language)
bp_depth = cursor.fetchall()

ws = wb['depthchart_bp']

#  Write Data
write_data(bp_depth)

#  Apply white text formatting to Columns O - X
row_limit = len(bp_depth) + 2
white_text = "O2:X" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "N" + str(len(bp_depth)+1)
borders(top_left,bottom_right)

########################
#  Get MI Depth Chart  #
########################

cursor.execute(dp.mi + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_mi + dp.order_language)
mi_depth = cursor.fetchall()

ws = wb['depthchart_MI']

#  Write Data
write_data(mi_depth)

#  Apply white text formatting to Columns N - AH
row_limit = len(mi_depth) + 2
white_text = "N2:AH" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(mi_depth)+1)
borders(top_left,bottom_right)

########################
#  Get 3B Depth Chart  #
########################

cursor.execute(dp.third_base + str(small_db.game_year) + dp.org_language + org_param + dp.pos_language_3b + dp.order_language)
third_base_depth = cursor.fetchall()

ws = wb['depthchart_3B']

#  Write Data
write_data(third_base_depth)

#  Apply white text formatting to Columns N - AG
row_limit = len(third_base_depth) + 2
white_text = "N2:AG" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(third_base_depth)+1)
borders(top_left,bottom_right)


# TEST SECTION
print("TEST SECTION:")
print(small_db.game_year)
print(small_db.game_date)
print("League ID: " + str(league_id))
print("Org Param: " + str(org_param))
print("Team Name: " + team_name)
print("str_date: " + str_date)
print("new file name: " + new_file_name)
print(first_base_depth)
print(wb.sheetnames)

# Close Everything Out
cnx.close()
remove('small_db')
wb.save(dest_path)
wb.close()