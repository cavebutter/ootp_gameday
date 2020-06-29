import xml.etree.ElementTree as ET
import mysql_db_config
import mysql.connector
from prettytable import PrettyTable
from pathlib import Path
from datetime import timedelta
from openpyxl import Workbook, worksheet
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle

#####################################
#  Formatting Styles and Functions  #
#####################################

# Table Headers
tableHeader = NamedStyle(name="tableHeader")
tableHeader.font = Font(bold=True, size=12)
tableHeader.alignment = Alignment(horizontal='center')
tbl_head_border = Side(style='medium')
tableHeader.border = Border(left=tbl_head_border,right=tbl_head_border,top=tbl_head_border,bottom=tbl_head_border)

def headers(cell_range):
    for _row in ws[cell_range]:
        for _cell in _row:
            _cell.style = tableHeader

# Table Body
#  Apply this style BEFORE headers
tableBody = NamedStyle(name="tableBody")
tableBody.font = Font(size=11)
bd = Side(style='thin')
tableBody.border = Border(left=bd,right=bd,top=bd,bottom=bd)

def body(cell_range):
    for _row in ws[cell_range]:
        for _cell in _row:
            _cell.style = tableBody

#  Inner Table Alignment
def inner_table(cell_range):
    for _row in ws[cell_range]:
        for _cell in _row:
           _cell.alignment = Alignment(horizontal='center')

#  Sheet Headers
sheet_header = NamedStyle(name="sheet_header")
sheet_header.font = Font(bold=True,size=14)

#  Section Headers
section_header = NamedStyle(name="section_header")
section_header.font = Font(bold=True,size=12)


#################
# Connect to db #
#################

database = input("What is the name of your OOTP database? ")
config = mysql_db_config.read_db_config(database)

try:
    mydb = mysql.connector.connect(**config)
    mycursor = mydb.cursor()

except mysql.connector.Error as err:
    print("Something went wrong: {}".format(err))

##################
#  Choose League #
##################

mycursor.execute("SELECT league_id, name FROM leagues")
league_list = mycursor.fetchall()


# - Make a Pretty Table to display ID and League
x = PrettyTable()
x.field_names = ['ID', 'League']
for item in league_list:
    x.add_row(item)
print(x)

league_id = input("Enter the league ID of your choice: ")

###########################
# Get current league date #
###########################
mycursor.execute("SELECT max(curr_date) FROM leagues")
db_return = mycursor.fetchall()  # TODO Is there a more elegant way to do this?
db_date0 = db_return[0]  # This too
db_date = db_date0[0]  # This too

####################################################
# Get league start date  & Other Date Calculations #
####################################################

sql = "SELECT start_date from leagues WHERE league_id=" + league_id
mycursor.execute(sql)
start_date_return = mycursor.fetchone()
start_date = start_date_return[0]

# Get today's day number and range of game days
date_diff = db_date - start_date
todays_day = date_diff.days
day_range = []
i = todays_day
for i in range(todays_day,todays_day+8):
    day_range.append(i)
    i+=1

# Get string date values for filename and year for SQL queries
file_start = db_date.strftime("%m_%d_%Y")
file_end_dt = db_date + timedelta(days=7)
file_end_str = file_end_dt.strftime("%m_%d_%Y")
file_dates = file_start + "-" + file_end_str
db_year = db_date.year

# Get date strings from day range for tab names
date_list = []
for i in range(0,7):
    str_date = db_date.strftime("%b-%d")
    date_list.append(str_date)
    db_date = db_date + timedelta(days=1)
    i+=1

###############
# Choose team #
###############

sql = """SELECT team_id, CONCAT(name, ' ', nickname) AS team
         FROM teams
         WHERE league_id=""" + league_id
mycursor.execute(sql)
teams_list = mycursor.fetchall()

# Make a Pretty Table
x = PrettyTable()
x.field_names = ['ID', 'Team']
for item in teams_list:
    x.add_row(item)
print(x)

team_id = input("Enter the ID for your team: ")

# Get team name
sql = "SELECT concat(name, ' ', nickname) as team FROM teams WHERE team_id = " + team_id
mycursor.execute(sql)
result = mycursor.fetchone()
team_name = result[0]

#######################################
# Find the schedule file and parse it #
#######################################

file_location = input("Enter the path for the schedule file: ")
path = Path(r'C:\Users\cohenja\Documents\pyOOTP\ILN_BGY_G162_SL1_D2_T4_T4_SL2_D2_T4_T4.lsdl')  # TODO For testing.  Change back to file_location
root = ET.parse(path)

#  Filter out all elements that are out of day range.
games = root.find('.//GAMES')
for g in games.findall('.GAME'):
    if int(g.attrib['day']) not in day_range:
        games.remove(g)

# add team_id's opponent to opp_list  # TODO add a marker for home/away so it can be used in headers

opp_list = []

for g in games.findall('.//GAME'):
    if g.attrib['home'] == team_id:
        opp_list.append(g.attrib['away'])
    elif g.attrib['away'] == team_id:
        opp_list.append(g.attrib['home'])

#####################################################
#  Get probable starters for today and next 7 days: #
#####################################################

i = 0
starters_id = []
for i in range(0,7):
    sql = 'SELECT starter_' + str(i) + ' FROM `projected_starting_pitchers` WHERE team_id=' + opp_list[i]
    mycursor.execute(sql)
    result = mycursor.fetchone()
    starters_id.append(result[0])

##########################
#  Get list of opp names #
##########################

opp_names = []
for i in range(0,len(opp_list)):
    sql = "SELECT concat(`name`,' ',`nickname`) as team FROM teams WHERE team_id = " + opp_list[i]
    mycursor.execute(sql)
    result = mycursor.fetchone()
    opp_names.append(result[0])

##################################
# Get names of starting pitchers #
##################################

starters_names = []
for i in range(0,len(starters_id)):
    sql = "SELECT concat(`first_name`,' ',`last_name`) as player FROM players WHERE player_id = " + str(starters_id[i])
    mycursor.execute(sql)
    result = mycursor.fetchone()
    starters_names.append(result[0])

###################################
# Get starting pitcher handedness # # TODO this is a total kluge.  I didn't want to complicate the zip below, but this is inelegant AF.
###################################
starter_throws = []
for i in range(0, len(starters_id)):
    sql = """SELECT IF(throws = 1, "RHP", "LHP") as throws
             FROM players WHERE player_id=""" + str(starters_id[i]) # TODO not sure why I can't make CASE work here.  Worth fixing?
    mycursor.execute(sql)
    result = mycursor.fetchone()
    starter_throws.append(result[0])

########################
# Zip up all the lists #
########################
iter_data = list(zip(opp_list,opp_names,starters_id,starters_names,date_list,starter_throws))

#########################
#  Excel Workbook Stuff #
#########################

wb = Workbook()
# Worksheet Styles
wb.add_named_style(tableBody)
wb.add_named_style(tableHeader)
wb.add_named_style(sheet_header)

############################
# Create tabs for each day #
############################

for i in range(0,len(iter_data)):

    opp_team_id = iter_data[i][0]
    opp_team_name = iter_data[i][1]
    starter_id = iter_data[i][2]
    starter_name = iter_data[i][3]
    date = iter_data[i][4]
    throws = iter_data[i][5]
    sheet_name = date
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    ######################################
    # Get active roster stats vs starter #
    ######################################
    stats_vs_starter_sql = """SELECT pos.pos_name as 'pos'
                , IF(players.bats = 1, 'R', IF(players.bats = 2, 'L', 'S')) as B
                , concat(players.first_name, ' ', players.last_name) as Player
                , pib.ab
                , pib.h
                , ROUND(pib.h / pib.ab, 3) as BA
                , pib.hr
            FROM players INNER JOIN players_individual_batting_stats as pib ON players.player_id = pib.player_id
            INNER JOIN players_roster_status as prs ON players.player_id = prs.player_id
            INNER JOIN positions as pos ON players.position=pos.position
            WHERE players.position <> 1 AND prs.team_id = """ + str(team_id) + ' AND prs.is_active = 1 AND pib.opponent_id =' + str(starter_id)
    mycursor.execute(stats_vs_starter_sql)
    stats = mycursor.fetchall()

    # Headers

    A1 = ws["A1"]
    A1.value = "Vs " + opp_team_name
    A1.style = sheet_header
    A3 = ws["A3"]
    A3.value = "Probable Starter - " + throws + " " + starter_name
    A3.style = sheet_header
    A5 = ws["A5"]
    A5.value = team_name + " Career Vs. " + throws + " " + starter_name
    A5.style = section_header
    G5 = ws['I5']
    G5.value = team_name + " Career Vs. " + opp_team_name + " Bullpen"
    G5.style = section_header

    #  Write the vs. Pitcher Headers
    vs_pitcher_headers = ["Pos", "B", "Player", "AB", "H", "Avg", "HR"]
    for k, val in enumerate(vs_pitcher_headers):
        ws.cell(row=7, column=k + 2).value = val

    # Write stats data to table
    for i, line in enumerate(stats):
        for k, val in enumerate(line):
            ws.cell(row=i + 8, column=k + 2).value = val

    # Format the stats table
    row_limit = len(stats) + 7
    inner_table_range = "C8:H" + str(row_limit)
    cell_range = "B8:H" + str(row_limit)
    body(cell_range)

    # Apply header formatting after body
    headers("B7:H7")

    # Format inner table fields
    inner_table(cell_range)

    ########################################
    # Get Active Players Stats vs. Bullpen #
    ########################################

    # Get Names and ID's of all Pitchers in Opp Bullpen
    bullpen_sql = """SELECT players.player_id, 
                        concat(players.first_name, ' ', players.last_name),
                        IF(players.throws = 1, 'RHP', 'LHP') AS throws
                    FROM players INNER JOIN players_roster_status as prs
                        ON players.player_id = prs.player_id
                    WHERE players.role IN (12,13) AND prs.is_active = 1 AND 
                    prs.team_id = """ + opp_team_id + " ORDER BY players.last_name"
    mycursor.execute(bullpen_sql)
    bullpen_roster = mycursor.fetchall()

    # iterate thru bullpen roster to get active roster stats vs. each reliever
    # counter to get the right row placement
    z = 6
    for reliever in bullpen_roster:
        rel_id = reliever[0]
        rel_name = reliever[1]
        throws = reliever[2]
        vs_rel_sql = """SELECT pos.pos_name as 'pos'
                        , IF(players.bats = 1, 'R', IF(players.bats = 2, 'L', 'S')) as B
                        , concat(players.first_name, ' ', players.last_name) as Player
                        , pib.ab
                        , pib.h
                        , ROUND(pib.h / pib.ab, 3) as BA
                        , pib.hr
                    FROM players INNER JOIN players_individual_batting_stats as pib ON players.player_id = pib.player_id
                    INNER JOIN players_roster_status as prs ON players.player_id = prs.player_id
                    INNER JOIN positions as pos ON players.position=pos.position
                    WHERE players.position <> 1 AND prs.team_id = """ + team_id + ' AND prs.is_active = 1 AND pib.opponent_id =' + str(rel_id)
        mycursor.execute(vs_rel_sql)
        batter_stats = mycursor.fetchall()

        # Write the data with header rows indicating the reliever
        ws.cell(row=z, column=10).value = throws + " " + rel_name
        ws.cell(row=z, column=10).style = section_header
        for k, val in enumerate(vs_pitcher_headers):
            ws.cell(row=z+1, column=k + 11).value = val
            ws.cell(row=z + 1, column=k + 11).style = tableHeader
        for i, line in enumerate(batter_stats):
            for k, val in enumerate(line):
                ws.cell(row=z + 2 + i, column=k + 11).value = val
                ws.cell(row=z + 2 + i, column=k + 11).style = tableBody
                ws.cell(row=z + 2 + i, column=k + 11).alignment = Alignment(horizontal='center')
        z += 3 + len(batter_stats)  # +1 for the table header, + 1 for new line + 1 for luck

    # Size Name columns
    ws.column_dimensions['D'].width = 19.00
    ws.column_dimensions['M'].width = 19.00
    ws.column_dimensions['B'].width = 5.00 #pos
    ws.column_dimensions['K'].width = 5.00 #pos
    ws.column_dimensions['C'].width = 3.00 # bats
    ws.column_dimensions['L'].width = 3.00 # bats
    # Remove Gridlines
    ws.sheet_view.showGridLines = False

#############################
#  Create Splits Worksheet  #
#############################
wb.create_sheet("Batting Splits")
ws = wb["Batting Splits"]
split_head = ws['A1']
split_head.style = sheet_header
split_head.value = team_name + " Active Roster Splits For " + str(db_year)
splits_header = ['pos', 'B', 'Player','PA','BB%','K%','OBP+','wOBA','wRC+']

# Left Split Stats
split_l = """SELECT pos.pos_name as 'pos'
                , IF(players.bats = 1, 'R', IF(players.bats = 2, 'L', 'S')) as B
                , CONCAT(players.first_name, ' ', players.last_name) AS Player 
                , l.PA
                , l.bbrate
                , l.krate
                , l.OBPplus
                , l.woba
                , l.`wRC+` 
            FROM players INNER JOIN CalcBatting_L AS l ON players.player_id=l.player_id
            INNER JOIN players_roster_status AS prs ON players.player_id = prs.player_id AND l.team_id = prs.team_id
            INNER JOIN positions pos ON players.position = pos.position
            WHERE players.position <> 1 AND year = """ + str(db_year) + " AND prs.team_id = " + str(team_id) + " AND prs.is_active = 1"

mycursor.execute(split_l)
left_split = mycursor.fetchall()

# Section Header
lhp_section = ws['B3']
lhp_section.value = "Splits Vs LHP"
lhp_section.style = section_header
rhp_section = ws['L3']
rhp_section.value = "Splits Vs RHP"
rhp_section.style = section_header

# Write header row
for k, val in enumerate(splits_header):
    ws.cell(row=4, column=k+2).value = val

# Write l_split data to table
for i, line in enumerate(left_split):
    for k, val in enumerate(line):
        ws.cell(row=i + 5, column=k + 2).value = val
        ws.cell(row=i + 5, column=6).style = 'Percent'  # TODO this is not working.
        ws.cell(row=i + 5, column=7).style = 'Percent'

# Formatting Stuff
row_limit = len(left_split) + 5
cell_range = "B4:J" + str(row_limit)
inner_table_range = "B5:J" + str(row_limit)
body(cell_range)
headers("B4:J4")
inner_table(inner_table_range)

# Right Split Stats
split_r = """SELECT pos.pos_name as 'pos'
                   , IF(players.bats = 1, 'R', IF(players.bats = 2, 'L', 'S')) as B
                   , CONCAT(players.first_name, ' ', players.last_name) AS Player 
                   , l.PA
                   , l.bbrate
                   , l.krate
                   , l.OBPplus
                   , l.woba
                   , l.`wRC+` 
               FROM players INNER JOIN CalcBatting_R AS l ON players.player_id=l.player_id
               INNER JOIN players_roster_status AS prs ON players.player_id = prs.player_id AND l.team_id = prs.team_id
               INNER JOIN positions as pos ON players.position=pos.position
               WHERE players.position <> 1 AND year = """ + str(db_year) + " AND prs.team_id = " + str(team_id) + " AND prs.is_active = 1"

mycursor.execute(split_r)
right_split = mycursor.fetchall()

# Write vsr header row
for k, val in enumerate(splits_header):
    ws.cell(row=4, column=k+12).value = val

# Write r_split data to table
for i, line in enumerate(right_split):
    for k, val in enumerate(line):
        ws.cell(row=i + 5, column=k + 12).value = val
        ws.cell(row=i + 5, column=16).style = 'Percent'
        ws.cell(row=i + 5, column=17).style = 'Percent'
# Formatting Stuff
row_limit = len(right_split) + 4
cell_range = "L4:T" + str(row_limit)
inner_table_range = "L5:T" + str(row_limit)
body(cell_range)
headers("L4:T4")
inner_table(inner_table_range)

# Remove Gridlines
ws.sheet_view.showGridLines = False

# Column Widths
ws.column_dimensions['D'].width = 19.00 #player
ws.column_dimensions['N'].width = 19.00 #player
ws.column_dimensions['B'].width = 5.00 # pos
ws.column_dimensions['L'].width = 5.00 # pos
ws.column_dimensions['M'].width = 3.00 # bats
ws.column_dimensions['C'].width = 3.00 # bats

del wb['Sheet']
save_path = Path.cwd() / 'output'
game_day_file_name = team_name + "_" + file_dates + '.xlsx'
wb.save(save_path / game_day_file_name)
mydb.close()
wb.close()
