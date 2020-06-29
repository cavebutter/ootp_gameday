import mysql.connector
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
import depth_chart_functions as dp
from prettytable import PrettyTable
import shutil
from pathlib import Path
import mysql_db_config


# Create Depth Charts

#########################
# Formatting Functions  #
#########################

#  Change font color to white
#  white_text param is the cell range
def white_font(white_text):
    ft = Font(color='ffffff')
    for _row in ws[white_text]:
        for _cell in _row:
            _cell.font = ft

#  Borders and Alignment
def borders(top_left, bottom_right):
    cell_range = top_left + ":" + bottom_right
    for _row in ws[cell_range]:
        for _cell in _row:
            _cell.border = Border(top=Side(border_style="thin"),
                                  left=Side(border_style="thin"),
                                  right=Side(border_style="thin"),
                                  bottom=Side(border_style="thin")
                                  )
    for _row in ws[cell_range]:  #  Is this second loop necessary?  Can I just add it to the one above?
        for _cell in _row:
            _cell.alignment = Alignment(horizontal='center')

#############################
#  Write to Excel Functions #
#############################

def write_data(depth_chart):
    for i, line in enumerate(depth_chart):
        for k, val in enumerate(line):
            ws.cell(row=i + 2, column=k + 1).value = val

#################
# Connect to DB #
#################

section = input("What is the name of your OOTP database?" )

config = mysql_db_config.read_db_config(section)

try:
    mydb = mysql.connector.connect(**config)
    mycursor = mydb.cursor()

except mysql.connector.Error as err:
    print("Something went wrong: {}".format(err))

# Get current league date
mycursor.execute("SELECT max(curr_date) FROM leagues")
db_return = mycursor.fetchall()  # Is there a more elegant way to do this?
db_date0 = db_return[0]  # This too
db_date = db_date0[0]  # This too
print("Current League Date: " + str(db_date))
year_param = db_date.year

#  Select League
mycursor.execute("SELECT league_id, name FROM leagues WHERE parent_league_id = 0")

league_list = mycursor.fetchall()


# - Make a Pretty Table to display ID and League
x = PrettyTable()
x.field_names = ['ID', 'League']
for item in league_list:
    x.add_row(item)
print(x)

league_id = input("Enter the league ID of your choice: ")

#  Get list of parent teams to select from
mycursor.execute("""SELECT team_id, CONCAT(name, ' ', nickname) AS team
                    FROM teams
                    WHERE team_id = parent_team_id AND league_id=""" + league_id)
teams_list = mycursor.fetchall()

# Make a Pretty Table
x = PrettyTable()
x.field_names = ['ID', 'Team']
for item in teams_list:
    x.add_row(item)
print(x)

#  Get org param
org_param = input("Enter the ID of the organization you want for your depth chart: ")

# Get org name
mycursor.execute("SELECT CONCAT(name, ' ', nickname) FROM teams WHERE team_id=" + org_param)
result = mycursor.fetchall()
team_name = result[0][0]

# Create New xlsx with the Date and Org Name
str_date = str(db_date)
new_file_name = team_name + "-" + str_date + '_depth_chart.xlsx'
src_file = Path.cwd() / 'depth_chart_template.xlsx'
dest_path = Path.cwd() / 'output' / new_file_name
shutil.copy(src_file, dest_path)

########################
#  Get 1b Depth Chart  #
########################

mycursor.execute(dp.first_base + str(year_param) + dp.org_language + org_param + dp.pos_language_1b + dp.order_language)
first_base_depth = mycursor.fetchall()

#  Open xlsx and find the correct tab
wb = openpyxl.load_workbook(dest_path)
ws = wb['depthchart_1B']

#  Write data beginning at Row 2

write_data(first_base_depth)

#  Apply white text formatting to Columns N - AG
row_limit = len(first_base_depth) + 2
white_text = "N2:AG" + str(row_limit)
white_font(white_text)

#  Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(first_base_depth)+1)
borders(top_left,bottom_right)

########################
#  Get OF Depth Chart  #
########################

mycursor.execute(dp.outfield + str(year_param) + dp.org_language + org_param + dp.pos_language_of + dp.order_language)
of_depth = mycursor.fetchall()

ws = wb['depthchart_of']

#  Write Data
write_data(of_depth)

#  Apply white text formatting to Columns N - AH
row_limit = len(of_depth) + 2
white_text = "N2:AH" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(of_depth)+1)
borders(top_left,bottom_right)

########################
#  Get C Depth Chart  #
########################

mycursor.execute(dp.catcher + str(year_param) + dp.org_language + org_param + dp.pos_language_c + dp.order_language)
c_depth = mycursor.fetchall()

ws = wb['depthchart_c']

#  Write Data
write_data(c_depth)

#  Apply white text formatting to Columns N - AF
row_limit = len(c_depth) + 2
white_text = "N2:AF" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(c_depth)+1)
borders(top_left,bottom_right)

########################
#  Get SP Depth Chart  #
########################

mycursor.execute(dp.sp + str(year_param) + dp.org_language + org_param + dp.pos_language_sp + dp.order_language)
sp_depth = mycursor.fetchall()

ws = wb['depthchart_sp']

#  Write Data
write_data(sp_depth)

#  Apply white text formatting to Columns O - X
row_limit = len(sp_depth) + 2
white_text = "O2:X" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "N" + str(len(sp_depth)+1)
borders(top_left,bottom_right)

########################
#  Get BP Depth Chart  #
########################

mycursor.execute(dp.sp + str(year_param) + dp.org_language + org_param + dp.pos_language_bp + dp.order_language)
bp_depth = mycursor.fetchall()

ws = wb['depthchart_bp']

#  Write Data
write_data(bp_depth)

#  Apply white text formatting to Columns O - X
row_limit = len(bp_depth) + 2
white_text = "O2:X" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "N" + str(len(bp_depth)+1)
borders(top_left,bottom_right)

########################
#  Get MI Depth Chart  #
########################

mycursor.execute(dp.mi + str(year_param) + dp.org_language + org_param + dp.pos_language_mi + dp.order_language)
mi_depth = mycursor.fetchall()

ws = wb['depthchart_MI']

#  Write Data
write_data(mi_depth)

#  Apply white text formatting to Columns N - AH
row_limit = len(mi_depth) + 2
white_text = "N2:AH" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(mi_depth)+1)
borders(top_left,bottom_right)

########################
#  Get 3B Depth Chart  #
########################

mycursor.execute(dp.third_base + str(year_param) + dp.org_language + org_param + dp.pos_language_3b + dp.order_language)
third_base_depth = mycursor.fetchall()

ws = wb['depthchart_3B']

#  Write Data
write_data(third_base_depth)

#  Apply white text formatting to Columns N - AG
row_limit = len(third_base_depth) + 2
white_text = "N2:AG" + str(row_limit)
white_font(white_text)

# Apply bordering to data
top_left = "A2"
bottom_right = "M" + str(len(third_base_depth)+1)
borders(top_left,bottom_right)

mydb.close()
wb.save(dest_path)
wb.close()
